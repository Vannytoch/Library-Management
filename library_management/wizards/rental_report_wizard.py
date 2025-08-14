from odoo import models, fields, api
from odoo.exceptions import UserError
from werkzeug.urls import url_encode
from datetime import  datetime
from dateutil.relativedelta import relativedelta
import base64,openpyxl,io
from openpyxl.styles import Alignment


class RentalReportWizard(models.TransientModel):
    _name = 'library.rental.report.wizard'
    _description = 'Library Rental Report Wizard'

    start_date = fields.Date(string="Start Date")
    end_date = fields.Date(string="End Date")
    add_data = fields.Binary(string="Import Data")
    file_name = fields.Char(string="File Name")
    export_file = fields.Binary(string='Download File', readonly=True)
    export_file_name = fields.Char(string="File Name")
    count_data = fields.Integer(string="Data count" ,compute="_compute_data_count")

    # status
    is_draft = fields.Boolean(string="Draft")
    is_confirmed = fields.Boolean(string="Confirmed")
    is_active = fields.Boolean(string="Active")
    is_returned = fields.Boolean(string="Returned")
    is_overdue = fields.Boolean(string="Overdue")

    def onchange_status(self):
        state = []
        for rec in self:
            if rec.is_draft:
                state.append('draft')
            else:
                if 'draft' in state:
                    state.remove('draft')

            if rec.is_confirmed:
                state.append('confirmed')
            else:
                if 'confirmed' in state:
                    state.remove('confirmed')

            if rec.is_active:
                state.append('active')
            else:
                if 'active' in state:
                    state.remove('active')

            if rec.is_returned:
                state.append('returned')
            else:
                if 'returned' in state:
                    state.remove('returned')

            if rec.is_overdue:
                state.append('overdue')
            else:
                if 'overdue' in state:
                    state.remove('overdue')
        return state

    @api.depends('start_date', 'end_date', 'is_draft','is_confirmed' ,'is_active' ,'is_returned', 'is_overdue')
    def _compute_data_count(self):
        for rec in self:
            if rec.start_date and rec.end_date:
                status = self.onchange_status()
                if len(status)>0:
                    data = self.env['library.rental'].search_count([
                        ('rental_date', '>=', rec.start_date),
                        ('rental_date', '<=', rec.end_date),
                        ('state', 'in', status)
                    ])
                else:
                    data = self.env['library.rental'].search_count([
                        ('rental_date', '>=', rec.start_date),
                        ('rental_date', '<=', rec.end_date)
                    ])
                rec.count_data = data
            else:
                rec.count_data = 0

    def action_import_data(self):
        for wizard in self:
            if not wizard.add_data:
                raise UserError("Please upload a file.")

            file_name = (wizard.file_name or "").lower()
            file_content = base64.b64decode(wizard.add_data)

            if file_name.endswith('.csv'):
                import csv
                from io import StringIO
                # Handle CSV
                csv_data = csv.reader(StringIO(file_content.decode('utf-8')))
                    # Process CSV row here
                    # create record logic...

            elif file_name.endswith('.xlsx'):
                import openpyxl
                from io import BytesIO
                # Handle Excel .xlsx
                workbook = openpyxl.load_workbook(filename=BytesIO(file_content))
                sheet = workbook.active

                last_rental = False  # store the last created rental
                for row in sheet.iter_rows(min_row=2, values_only=True):  # skip header
                    # Check if row[0] to row[5] are all empty
                    if all(cell in (None, '') for cell in row[0:6]):
                        # Blank row → add books to last rental only
                        if not last_rental:
                            continue  # skip if no rental exists yet
                        book_titles_raw = row[6]  # or some column that contains books
                        # Convert titles to rental lines
                        rental_lines = []
                        if book_titles_raw:
                            book_titles = [title.strip() for title in str(book_titles_raw).split(',')]
                            for title in book_titles:
                                book = self.env['library.book'].search([('title', '=', title)], limit=1)
                                if book:
                                    rental_lines.append((0, 0, {'book_id': book.id}))
                                else:
                                    raise UserError(f"Book with title '{title}' not found.")

                        # Add lines to the last rental
                        if rental_lines:
                            last_rental.write({'rental_line_ids': rental_lines})
                        continue  # skip creating a new rental

                    book_titles_raw = row[0]

                    due_date = row[0]
                    member = row[1]
                    rental_date = row[2]
                    rental_fee = row[3]
                    return_date = row[4]
                    status = row[5]
                    title = row[6]
                    qty = row[7]
                    discount = row[8]
                    member_id =0

                    if member:
                        member_id = self.env['library.member'].search([('name', '=', member)], limit=1)
                    # Convert titles to book IDs
                    rental_line = []
                    book = self.env['library.book'].search([('title', '=', title)], limit=1)
                    if book:
                        rental_line.append((0, 0,
                                            {'book_id': book.id,
                                             'discount': float(discount or 0),
                                             'qty': int(qty or 0)}
                                            ))
                    else:
                        raise UserError(f"Book with title '{title}' not found.")

                    # Create rental record
                    last_rental = self.env['library.rental'].create({
                        'member_id': member_id.id,
                        'rental_line_ids': rental_line,
                        'rental_date': rental_date,
                        'due_date': due_date,
                        'return_date': return_date,
                        'total_amount': rental_fee,
                        'state': status if status in ['draft', 'returned'] else 'draft',
                    })

            else:
                raise UserError("Unsupported file format. Please upload CSV or XLSX file.")
            # ✅ return action to keep the wizard open

    def action_export_data(self):
        # Generate URL with params
        base_url = '/library/export_rental_xlsx?'
        # Start building query parameters
        query_params = {}
        # Add start and end date if they exist
        if self.start_date:
            query_params['start_date'] = str(self.start_date)
        if self.end_date:
            query_params['end_date'] = str(self.end_date)
        # Add state only if onchange_status() returns something
        state_list = self.onchange_status()
        if state_list:
            query_params['state'] = ','.join(state_list)
        # Now encode the final query string
        query = url_encode(query_params)
        full_url = base_url + query

        return {
            'type': 'ir.actions.act_url',
            'url': full_url,
            'target': 'self',
        }


    def action_generate_report(self):
        if self.count_data <1:
            raise UserError("No data to generate!")
        data = {
            'form': {
                'start_date': self.start_date,
                'end_date': self.end_date,
            }
        }
        return self.env.ref('library_management.action_rental_report_pdf').report_action(self, data=data)

    def generate_and_send_report(self):
        # Fetch records
        start_date = datetime.now() - relativedelta(months=1)
        end_date = datetime.now()
        domain = [
            ('rental_date', '>=', start_date),
            ('rental_date', '<=', end_date)
        ]
        rentals = self.env['library.rental'].sudo().search(domain)

        # Create XLSX
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Rental Report'
        ws.append(['Book', 'Due Date', 'Member', 'Rental Date', 'Rental Fee', 'Return Date', 'Status'])

        ws.column_dimensions['A'].width = 30  # Book Title
        ws.column_dimensions['B'].width = 30  # Date
        ws.column_dimensions['C'].width = 30  # Customer
        ws.column_dimensions['D'].width = 30  # Start
        ws.column_dimensions['E'].width = 30  # Fee
        ws.column_dimensions['F'].width = 30  # Return
        ws.column_dimensions['G'].width = 30  # State
        for r in rentals:
            book_title = ', '.join(r.rental_line_ids.mapped('book_id.title'))
            ws.append([book_title, r.due_date, r.member_id.name, r.rental_date, r.total_amount, r.return_date, r.state])

        for row in ws.iter_rows(min_row=2):
            row[0].alignment = Alignment(wrap_text=True)

        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)
        xlsx_data = fp.read()

        # Send Email
        attachment = self.env['ir.attachment'].create({
            'name': f"rental_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
            'type': 'binary',
            'datas': base64.b64encode(xlsx_data),
            'res_model': 'library.rental',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Replace with your real recipient email
        recipient_email = 'admin@example.com'
        mail_values = {
            'subject': 'Monthly Rental Report',
            'body_html': '<p>Please find attached the monthly rental report.</p>',
            'email_to': recipient_email,
            'attachment_ids': [(6, 0, [attachment.id])],
        }
        self.env['mail.mail'].sudo().create(mail_values).send()